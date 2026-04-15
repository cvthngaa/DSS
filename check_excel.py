import openpyxl
wb = openpyxl.load_workbook('tsp_results.xlsx')
print("=" * 60)
print("SHEETS:", wb.sheetnames)
print("=" * 60)

for s in wb.sheetnames:
    ws = wb[s]
    print(f"\n--- {s} ({ws.max_row} rows x {ws.max_column} cols) ---")
    for row in ws.iter_rows(max_row=min(ws.max_row, 12), values_only=True):
        print(row)
